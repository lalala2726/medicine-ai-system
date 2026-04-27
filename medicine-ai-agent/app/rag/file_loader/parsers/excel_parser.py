from __future__ import annotations

import csv
from pathlib import Path

from app.core.exception.exceptions import ServiceException
from app.rag.file_loader.parsers.base import BaseParser


# ---------------------------------------------------------------------------
# 原始行数据提取（供 Excel 行切片使用）
# ---------------------------------------------------------------------------

def _extract_rows_xlsx(file_path: Path) -> list[list[str]]:
    """从 xlsx 文件提取原始行数据。"""
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ServiceException("缺少 openpyxl 依赖，无法解析 xlsx 文件") from exc

    workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    rows: list[list[str]] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            if any(v.strip() for v in values):
                rows.append(values)
    return rows


def _extract_rows_xls(file_path: Path) -> list[list[str]]:
    """从 xls 文件提取原始行数据。"""
    try:
        import xlrd
    except Exception as exc:
        raise ServiceException("缺少 xlrd 依赖，无法解析 xls 文件") from exc

    workbook = xlrd.open_workbook(str(file_path))
    rows: list[list[str]] = []
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        for row_index in range(sheet.nrows):
            values = [
                str(sheet.cell_value(row_index, col_index))
                for col_index in range(sheet.ncols)
            ]
            if any(v.strip() for v in values):
                rows.append(values)
    return rows


def _extract_rows_csv(file_path: Path) -> list[list[str]]:
    """从 csv 文件提取原始行数据。"""
    rows: list[list[str]] = []
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            values = [str(cell) for cell in row]
            if any(v.strip() for v in values):
                rows.append(values)
    return rows


def parse_rows(file_path: Path) -> list[list[str]]:
    """提取 Excel / CSV 文件的原始行数据（二维字符串列表）。

    Args:
        file_path: 文件路径，支持 .xlsx / .xls / .csv。

    Returns:
        行列表，每行为单元格字符串列表。

    Raises:
        ServiceException: 不支持的文件格式或依赖缺失。
    """
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _extract_rows_xlsx(file_path)
    if suffix == ".xls":
        return _extract_rows_xls(file_path)
    if suffix == ".csv":
        return _extract_rows_csv(file_path)
    raise ServiceException(f"不支持的 Excel/CSV 格式: {suffix}")


# ---------------------------------------------------------------------------
# 文本提取（兼容旧流程）
# ---------------------------------------------------------------------------

def _parse_xlsx(file_path: Path) -> str:
    """解析 xlsx 文件并拼接为单一文本，保留工作表名称作为结构标签。"""
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ServiceException("缺少 openpyxl 依赖，无法解析 xlsx 文件") from exc

    workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    sections: list[str] = []
    for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
        sheet = workbook[sheet_name]
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            if any(value.strip() for value in values):
                rows.append("\t".join(values))
        sheet_text = "\n".join(rows).strip()
        if sheet_text:
            sections.append(f"Sheet {sheet_index}: {sheet_name}\n{sheet_text}")
    return "\n\n".join(sections)


def _parse_xls(file_path: Path) -> str:
    """解析 xls 文件并拼接为单一文本，保留工作表名称作为结构标签。"""
    try:
        import xlrd
    except Exception as exc:
        raise ServiceException("缺少 xlrd 依赖，无法解析 xls 文件") from exc

    workbook = xlrd.open_workbook(str(file_path))
    sections: list[str] = []
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        rows: list[str] = []
        for row_index in range(sheet.nrows):
            values = [
                str(sheet.cell_value(row_index, col_index))
                for col_index in range(sheet.ncols)
            ]
            if any(value.strip() for value in values):
                rows.append("\t".join(values))
        sheet_text = "\n".join(rows).strip()
        if sheet_text:
            sections.append(
                f"Sheet {sheet_index + 1}: {sheet.name}\n{sheet_text}"
            )
    return "\n\n".join(sections)


def _parse_csv(file_path: Path) -> str:
    """解析 csv 文件并拼接为单一文本。"""
    rows: list[str] = []
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            values = [str(cell) for cell in row]
            if any(v.strip() for v in values):
                rows.append("\t".join(values))
    return "\n".join(rows)


class ExcelParser(BaseParser):
    """解析 Excel / CSV 文件，支持 xlsx / xls / csv。"""

    def parse(self, file_path: Path) -> str:
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            return _parse_xlsx(file_path)
        if suffix == ".xls":
            return _parse_xls(file_path)
        if suffix == ".csv":
            return _parse_csv(file_path)
        raise ServiceException(f"不支持的 Excel/CSV 格式: {suffix}")
